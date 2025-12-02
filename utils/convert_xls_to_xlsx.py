#!/usr/bin/env python3
"""
Utility to convert .xls files to .xlsx format
AI generated to solve problem with the Daystar timetable being a .xls file
"""
import xlrd
from openpyxl import Workbook
from pathlib import Path
import sys


def convert_xls_to_xlsx(xls_path, xlsx_path=None):
    """
    Convert an .xls file to .xlsx format

    Args:
        xls_path: Path to the input .xls file
        xlsx_path: Path to the output .xlsx file (optional, defaults to same name with .xlsx)

    Returns:
        Path to the converted .xlsx file
    """
    xls_path = Path(xls_path)

    if xlsx_path is None:
        xlsx_path = xls_path.with_suffix('.xlsx')
    else:
        xlsx_path = Path(xlsx_path)

    # Read the .xls file
    rb = xlrd.open_workbook(str(xls_path), formatting_info=False)

    # Create a new .xlsx workbook
    wb = Workbook()

    # Remove the default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Copy each sheet
    for sheet_index in range(rb.nsheets):
        rs = rb.sheet_by_index(sheet_index)
        ws = wb.create_sheet(title=rs.name)

        # Copy all cells
        for row_idx in range(rs.nrows):
            for col_idx in range(rs.ncols):
                cell_value = rs.cell_value(row_idx, col_idx)
                cell_type = rs.cell_type(row_idx, col_idx)

                # Handle different cell types
                if cell_type == xlrd.XL_CELL_DATE:
                    # Convert Excel date to Python datetime
                    date_tuple = xlrd.xldate_as_tuple(cell_value, rb.datemode)
                    from datetime import datetime
                    cell_value = datetime(*date_tuple)
                elif cell_type == xlrd.XL_CELL_EMPTY:
                    cell_value = None
                elif cell_type == xlrd.XL_CELL_BOOLEAN:
                    cell_value = bool(cell_value)

                # Write to the new workbook (openpyxl uses 1-based indexing)
                ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)

    # Save the .xlsx file
    wb.save(str(xlsx_path))
    print(f"Converted {xls_path} to {xlsx_path}")

    return xlsx_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python convert_xls_to_xlsx.py <input.xls> [output.xlsx]")
        sys.exit(1)

    xls_file = sys.argv[1]
    xlsx_file = sys.argv[2] if len(sys.argv) > 2 else None

    convert_xls_to_xlsx(xls_file, xlsx_file)


