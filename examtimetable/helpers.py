# contains helper functions
from openpyxl import load_workbook
from datetime import datetime
from typing import Any, Dict, List, Optional
import re
from utils.compute_datetime import compute_datetime_str

# parses nursing exam timetable
def nursing_exam_timetable_parser(file):

    # Define the list of column headers
    column_headers = ['Day', 'Campus', 'Coordinator', 'Courses', 'Hours',
                      'Venue', 'Invigilators', 'Courses_Afternoon',
                      'Hours_Afternoon', 'Invigilators_Afternoon',
                      'Venue_Afternoon']

    def extract_course_info(column_data_dict, time_key, time_range):
        courses = []
        existing_course_codes = set()
        for i in range(len(column_data_dict["Day"])):
            course_code = column_data_dict[time_key][i]
            if course_code not in time_range and course_code not in existing_course_codes:
                day_val = column_data_dict["Day"][i]
                time_val = '8:30AM-11:30AM' if "8.30" in time_range[0] else '1:30PM-4:30PM'

                course_info = {
                    "course_code": column_data_dict[time_key][i].strip(), #.replace(" ", ""),
                    "coordinator": column_data_dict["Coordinator"][i],
                    "time": time_val,
                    "day": day_val,
                    "campus": column_data_dict["Campus"][i],
                    "hrs": column_data_dict[f"Hours{'_Afternoon' if '_Afternoon' in time_key else ''}"][i],
                    "venue": column_data_dict[f"Venue{'_Afternoon' if '_Afternoon' in time_key else ''}"][i],
                    "invigilator": column_data_dict[f"Invigilators{'_Afternoon' if '_Afternoon' in time_key else ''}"][i],
                    "datetime_str": compute_datetime_str(day_val, time_val)
                }

                courses.append(course_info)
                existing_course_codes.add(course_code)
        return courses

    # Loading the excel workbook
    wb_obj = load_workbook(file)
    sheet = wb_obj.active  # Activating the sheet for use

    # Initialize dictionary to store column data
    column_data_dict = {}

    # Iterate over columns and store data in dictionary
    for i, column in enumerate(sheet.iter_cols(values_only=True) if sheet else []):
        last_value = None
        column_data = []
        for cell in column:
            if cell is not None:
                # if column_headers[i] == 'Day':
                #     cell = cell.strftime('%A %d-%m-%Y')
                last_value = cell
                column_data.append(cell)
            else:
                column_data.append(last_value)
        if any(cell is not None for cell in column_data):
            column_data_dict[column_headers[i]] = column_data

    # Extract course information
    morning_exams = extract_course_info(column_data_dict, "Courses", ['8.30AM-11.30AM', '8.30-11.30 AM'])
    afternoon_exams = extract_course_info(column_data_dict, "Courses_Afternoon", ['1.30PM-4.30PM', '1.30-4.30PM'])

    # Combine morning and afternoon exams
    courses = morning_exams + afternoon_exams

    return courses


# parses nursing school timetable
def parse_nursing_timetable(file_path):
    # holds start and end times of every column number
    time_dictionary = {
        "2": ("8AM", "9AM"),
        "3": ("9AM", "10AM"),
        "4": ("10AM", "11AM"),
        "5": ("11AM", "12PM"),
        "6": ("12PM", "1PM"),
        "7": ("1PM", "2PM"),
        "8": ("2PM", "3PM"),
        "9": ("3PM", "4PM"),
        "10": ("4PM", "5PM"),
        "11": ("5PM", "6PM"),
    }

    days_of_the_week = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"]
    unnecessary_course_values = ["CHAPEL", "CLP", "SDL", "KEY",
                                "CLP-CLINICAL PRACTICE",
                                "SDL- SELF DIRECTED LEARNING"]

    courses = [] # will hold dictionaries of courses

    # loading the excel workbook
    wb_obj = load_workbook(filename=file_path)
    work_sheets = wb_obj.sheetnames # getting available work sheets

    # getting info from from first worksheet
    first_work_sheet = wb_obj[work_sheets[0]]

    # will hold course names and lecturers key being course code
    course_lectures = {}

    # iterating through the sheet
    for row in first_work_sheet.iter_rows(values_only=True):
        if row[1] == "CODE" or row[1] is None:
            continue

        # concantenating course name and the lecturer
        course_lectures[row[1]] = [row[2] , row[3]]

     # getting info from from second worksheet
    second_work_sheet = wb_obj[work_sheets[1]]

    # course contents
    day = ""
    venue = ""
    course_time = ""
    course_name = ""

    # itearating the workbook
    for row_idx, rows in enumerate(second_work_sheet.iter_rows(values_only=True)):
        # skipping the unnecessary titles
        if row_idx in [0, 1, 2]:
            continue

        # skipping row with dates and taking the day of the course
        if rows[1] in days_of_the_week:
            day = rows[1]
            continue

        for idx ,row_value in enumerate(rows):
            # skipping the cohort part
            if idx == 1:
                continue

            # skipping unnecessary course values
            if row_value is None or (isinstance(row_value, str) and row_value.strip() in unnecessary_course_values):
                continue

            # getting the venue
            if idx == 0:
                venue = row_value
                continue

            # getting course time
            start_time = time_dictionary[f"{idx}"][0]

            # setting course code
            course_name = str(row_value).strip() if row_value is not None else ""

            # handling merged rows
            remaining_cells = rows[idx+1:]
            rem_idx = idx + 1
            for rem_idx_iter, rem in enumerate(remaining_cells, start=idx+1):
                if rem is not None:
                    rem_idx = rem_idx_iter
                    break

            # concatenating start time and end time
            course_time = start_time + "-" + time_dictionary[f"{rem_idx}"][1]

            courses.append({
                "course_code": course_name[:7].strip().replace(" ", ""),
                "lecturer": course_lectures[course_name[:7].strip()][1],
                "course_name": course_name,
                "day":day,
                "venue":venue,
                "time":course_time
                })
    print(courses)
    return courses

def parse_school_exam_timetable(file):
    def time_difference(start_time, end_time):
        """
        Returns the difference in hrs between two
        time intervals
        """
        start_time = start_time.strip()
        end_time = end_time.strip()

        formats = ["%I:%M%p", "%H:%M", "%I:%M %p", "%H:%M "]
        start_dt = None
        end_dt = None

        for fmt in formats:
            try:
                start_dt = datetime.strptime(start_time, fmt)
                break
            except ValueError:
                continue

        for fmt in formats:
            try:
                end_dt = datetime.strptime(end_time, fmt)
                break
            except ValueError:
                continue

        if start_dt is None or end_dt is None:
            return "2"

        hrs = (end_dt - start_dt).total_seconds() / 3600
        return str(hrs)

    # loading the workbook
    wb_obj = load_workbook(filename=file)

    work_sheets = wb_obj.sheetnames # getting available work sheets

    rooms = {} # will hold room informationhold information about courses

    courses = [] # will hold courses information

    days_of_the_week = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY"]
    for sheet in work_sheets:
        work_sheet = wb_obj[sheet]

        # iterating through the rooms and storing them
        for column_one in work_sheet.iter_cols(values_only=True): #For each Column in a worksheet
            for i, room in enumerate(column_one):
                # skipping none values and room word
                if room is None or room == "ROOM":
                    continue
                rooms[f'{i}'] = room
            break
        # a list of the remaining columns without the rooms
        data_columns = list(work_sheet.iter_cols(values_only=True))[1:]

        # values for the course code
        day = ""
        course_time = ""
        course_code = ""
        hours = "2"

        # iterating through the columns data
        for column in data_columns:
            for idx, value in enumerate(column):
                # skipping empty cell values
                if value is None:
                    continue

                if value == "CHAPEL":
                    continue

                # checking if its date and day specification
                if isinstance(value, datetime):
                    day = days_of_the_week[value.weekday()] + " " + str(value.date()).replace("-", "/")
                elif isinstance(value, str) and value.split(" ")[0] in days_of_the_week:
                    day = value

                # checking if its time specification
                elif isinstance(value, str) and len(value) > 0 and value[0].isdigit():
                    course_time = value.strip()
                    # print(course_time)
                    if "-" in course_time:
                        start_time = course_time.split("-")[0]
                        end_time = course_time.split("-")[1]
                        hours = time_difference(start_time, end_time)
                    else:
                        hours = "2"
                elif isinstance(value, str):
                    course_code = value
                    hours_str = "2"  # default
                    if hours and len(hours) > 0:
                        hours_str = "2" if hours[0] == "-" else str(hours)
                    courses.append(
                        {
                            "course_code": course_code,
                            "day": day,
                            "time": course_time,
                            "venue": rooms.get(f"{idx}", ""),
                            "campus": "",
                            "coordinator": "",
                            "hrs": hours_str,
                            "invigilator": "",
                            "datetime_str": None,
                        }
                    )
    return courses

def strath_extractor(file):
    """
    Extracts TT data for Strathmore University
    """

    def format_strath_time(time):
        """
        converts format (8:00-10:00) to (8:00AM-10:00AM)
        Handles edge cases like "11:00AM-14.00"
        """
        if not time or "-" not in time:
            return time

        original_time = str(time).strip()
        clean_time = original_time.upper()
        clean_time = re.sub(r'\s*-\s*', '-', clean_time)

        start_time, end_time = clean_time.split("-", 1)
        start_time = start_time.strip()
        end_time = end_time.strip()

        def convert_to_12hour(time_24, ampm_hint=None):
            """Convert 24-hour time to 12-hour format"""
            if not time_24:
                return time_24

            ampm = ampm_hint
            time_str = time_24

            ampm_match = re.search(r'([AP]M)', time_str)
            if ampm_match:
                ampm = ampm_match.group(1)
                time_str = re.sub(r'[AP]M', '', time_str).strip()

            if ":" in time_str:
                hour, minute = time_str.split(":")
                hour = int(hour)
                minute = int(minute)
            elif "." in time_str:
                hour, minute = time_str.split(".")
                hour = int(hour)
                minute = int(minute)
            elif len(time_str) == 4 and time_str.isdigit():
                hour = int(time_str[:2])
                minute = int(time_str[2:])
            else:
                try:
                    hour = int(time_str)
                    minute = 0
                except ValueError:
                    return time_24

            if not ampm:
                if hour >= 12:
                    ampm = 'PM'
                    if hour > 12:
                        hour = hour - 12
                else:
                    ampm = 'AM'
                    if hour == 0:
                        hour = 12

            return f"{hour}:{minute:02d}{ampm}"

        start_ampm = None
        end_ampm = None

        start_ampm_match = re.search(r'([AP]M)', start_time)
        if start_ampm_match:
            start_ampm = start_ampm_match.group(1)

        end_ampm_match = re.search(r'([AP]M)', end_time)
        if end_ampm_match:
            end_ampm = end_ampm_match.group(1)

        formatted_start_time = convert_to_12hour(start_time, start_ampm)
        formatted_end_time = convert_to_12hour(end_time, end_ampm or start_ampm)

        return f"{formatted_start_time}-{formatted_end_time}"

    web_obj = load_workbook(file)
    sheet = web_obj.active

    # Variables to track merged cells
    current_date = ""
    current_time = ""
    current_course = ""
    current_group = ""
    current_number = ""
    current_venue = ""
    current_lecturer = ""

    course = []

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True) if sheet else []):
        # Skip header rows (assuming first 3 rows are intro/header)
        if row_idx < 3:
            continue

        date_val = row[0] if len(row) > 0 else None
        time_val = row[2] if len(row) > 2 else None
        course_val = row[4] if len(row) > 4 else None
        group_val = row[6] if len(row) > 6 else None
        number_val = row[7] if len(row) > 7 else None
        venue_val = row[8] if len(row) > 8 else None
        lecturer_val = row[10] if len(row) > 10 else None

        # Handle merged cells - update current values when new data is found
        if date_val and str(date_val).strip():
            current_date = str(date_val).strip().rstrip('.')
        if time_val and str(time_val).strip():
            current_time = str(time_val).strip()
        if course_val and str(course_val).strip():
            current_course = str(course_val).strip()
        if group_val and str(group_val).strip():
            current_group = str(group_val).strip()
        if number_val is not None:  # Allow 0 as valid
            current_number = str(number_val).strip()
        if venue_val and str(venue_val).strip():
            current_venue = str(venue_val).strip()
        if lecturer_val and str(lecturer_val).strip():
            current_lecturer = str(lecturer_val).strip()

        # Append if there's a new group or a new venue (for split venues)
        if current_course and current_group and current_venue:
            if (group_val and str(group_val).strip()) or (venue_val and str(venue_val).strip()):
                # Split course code and name
                course_parts = current_course.split(":", 1)
                course_code = course_parts[0].strip() if course_parts else current_course
                course_name = course_parts[1].strip() if len(course_parts) > 1 else ""

                # Format time to match existing pattern
                formatted_time = format_strath_time(current_time)

                course_info = {
                    "course_code": course_code,
                    "course_name": course_name,
                    "day": current_date,
                    "time": formatted_time,
                    "venue": current_venue,
                    "lecturer": current_lecturer,
                    "group": current_group,
                    "student_count": current_number,
                    "program": current_group.split()[0] if current_group else "",
                }
                course.append(course_info)

    return course

def kca_extractor(file):
    """
    Extracts TT data for KCA University
    Returns standardized format: course_code, day, time, venue
    """

    def compute_hours(formatted_time: str) -> str:
        """
        Defaults to '2' on any parsing issues.
        """
        if not formatted_time or "-" not in str(formatted_time):
            return "2"

        try:
            start_str, end_str = str(formatted_time).split("-", 1)
            start_str = start_str.strip()
            end_str = end_str.strip()

            def parse_time(s: str):
                for fmt in ["%I:%M%p", "%I:%M %p", "%H:%M", "%H.%M"]:
                    try:
                        return datetime.strptime(s, fmt)
                    except ValueError:
                        continue
                return None

            start_dt = parse_time(start_str)
            end_dt = parse_time(end_str)
            if not start_dt or not end_dt:
                return "2"

            hrs = (end_dt - start_dt).total_seconds() / 3600
            return str(int(hrs)) if hrs.is_integer() else str(hrs)
        except Exception:
            return "2"

    def format_time(time_str):
        """
        Standardizes various time formats to (8:00AM-10:00AM)
        Handles: "8:00AM-10:00AM", "0800-1000 HRS", "2.30 pm - 4.30pm", "11:00AM-14.00"
        """
        if not time_str:
            return ""

        original_time = str(time_str).strip()
        clean_time = original_time.upper()

        clean_time = re.sub(r'(HR|HRS)', '', clean_time).strip()
        clean_time = re.sub(r'\s*-\s*', '-', clean_time)
        clean_time = re.sub(r'\s+', ' ', clean_time)

        def to_12hour(hour_min, ampm=None):
            """Convert hour:minute to 12-hour format"""
            try:
                if ':' in hour_min:
                    hour, minute = map(int, hour_min.split(':'))
                else:
                    hour = int(hour_min)
                    minute = 0

                if ampm is None:
                    if hour >= 12:
                        ampm = 'PM'
                        if hour > 12:
                            hour = hour - 12
                        elif hour == 12:
                            pass
                    else:
                        ampm = 'AM'
                        if hour == 0:
                            hour = 12

                return f"{hour}:{minute:02d}{ampm}"
            except (ValueError, AttributeError):
                return hour_min

        start_time_str = None
        end_time_str = None
        start_ampm = None
        end_ampm = None

        if '-' in clean_time:
            parts = clean_time.split('-', 1)
            start_part = parts[0].strip()
            end_part = parts[1].strip()

            start_ampm_match = re.search(r'([AP]M)', start_part)
            if start_ampm_match:
                start_ampm = start_ampm_match.group(1)
                start_part = re.sub(r'[AP]M', '', start_part).strip()

            end_ampm_match = re.search(r'([AP]M)', end_part)
            if end_ampm_match:
                end_ampm = end_ampm_match.group(1)
                end_part = re.sub(r'[AP]M', '', end_part).strip()

            start_time_str = start_part
            end_time_str = end_part

            if not start_time_str or not end_time_str:
                return original_time

            start_time_str = start_time_str.replace('.', ':')
            end_time_str = end_time_str.replace('.', ':')

            if ':' not in start_time_str and len(start_time_str) == 4 and start_time_str.isdigit():
                start_time_str = f"{start_time_str[:2]}:{start_time_str[2:]}"
            elif ':' not in start_time_str:
                start_time_str = f"{start_time_str}:00"

            if ':' not in end_time_str:
                if len(end_time_str) == 4 and end_time_str.isdigit():
                    end_time_str = f"{end_time_str[:2]}:{end_time_str[2:]}"
                elif '.' in end_time_str or end_time_str.replace('.', '').isdigit():
                    end_time_str = end_time_str.replace('.', ':')
                    if ':' not in end_time_str:
                        end_time_str = f"{end_time_str}:00"
                else:
                    end_time_str = f"{end_time_str}:00"

            formatted_start = to_12hour(start_time_str, start_ampm)
            formatted_end = to_12hour(end_time_str, end_ampm or start_ampm)

            return f"{formatted_start}-{formatted_end}"

        return original_time

    def convert_date(date_val):
        """Convert Excel serial or string to readable date
        from (Monday, 1st January 2025) to (2025-01-01)
        """
        if isinstance(date_val, (int, float)):
            try:
                return datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(date_val) - 2).strftime('%Y-%m-%d')
            except ValueError:
                return str(date_val)
        return str(date_val).strip()



    def normalize_header(header_str):
        if not header_str:
            return ""
        normalized = str(header_str).upper().strip()
        normalized = re.sub(r'[^\w\s]', '', normalized)
        normalized = re.sub(r'\s+', '_', normalized)
        return normalized

    wb_obj = load_workbook(file, data_only=True)

    key_map = {
        "DATE": "DATE",
        "TIME": "TIME",
        "VENUE": "ROOM|VENUE",
        "UNIT_CODE": "UNIT_CODE|UNIT CODE",
        "CAMPUS": "CAMPUS",
        "COORDINATOR": "INVIGILATOR OF THE DAY",
        "INVIGILATOR": "INVIGILATORS|PRINCIPAL INVIGILATORS|INVIGILATOR",
    }
    unit_code_pattern = re.compile(r"^[A-Z]{2,4}\s*\d{3,4}")

    all_courses: List[Dict[str, Any]] = []

    for sheet_name in wb_obj.sheetnames:
        sheet = wb_obj[sheet_name]

        header_row: Optional[List[str]] = None
        header_idx = 0
        for row_idx, row in enumerate(
            sheet.iter_rows(values_only=True) if sheet else [], start=1
        ):
            if any("UNIT CODE" in str(cell).upper() for cell in row if cell):
                header_row = [str(cell) if cell is not None else "" for cell in row]
                header_idx = row_idx
                break

        if not header_row:
            continue

        norm_headers = {
            normalize_header(h): idx
            for idx, h in enumerate(header_row)
            if h
        }

        courses: List[Dict[str, Any]] = []
        current_entry: Optional[Dict[str, Any]] = None

        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=header_idx + 1, values_only=True) if sheet else [],
            start=header_idx + 1,
        ):
            row = [cell if cell is not None else "" for cell in row]

            unit_code = ""
            for pattern in key_map["UNIT_CODE"].split("|"):
                norm_pattern = normalize_header(pattern)
                if norm_pattern in norm_headers:
                    unit_code = str(row[norm_headers[norm_pattern]]).strip()
                    break

            norm_uc = unit_code.upper().replace(" ", "").replace("_", "")
            if (
                not unit_code
                or norm_uc in {"UNITCODE", "UNIT", "TIME", "DATE", "DAY", "VENUE", "ROOM", "EXAMINATION"}
                or not unit_code_pattern.search(unit_code.upper())
            ):
                continue

            if current_entry:
                courses.append(current_entry)

            current_entry = {
                "course_code": unit_code,
                "day": "",
                "time": "",
                "venue": "",
                "campus": "",
                "coordinator": "",
                "hrs": "",
                "invigilator": "",
                "datetime_str": None,
            }

            for out_key, patterns in key_map.items():
                for pattern in patterns.split("|"):
                    norm_pattern = normalize_header(pattern)
                    if norm_pattern in norm_headers:
                        val = row[norm_headers[norm_pattern]]
                        if out_key == "DATE":
                            val = convert_date(val)
                            current_entry["day"] = str(val).strip()
                        elif out_key == "TIME":
                            raw_time = str(val)
                            formatted = format_time(raw_time)
                            current_entry["time"] = formatted
                            current_entry["hrs"] = compute_hours(formatted)
                        elif out_key == "VENUE":
                            current_entry["venue"] = str(val).strip()
                        elif out_key == "CAMPUS":
                            current_entry["campus"] = str(val).strip()
                        elif out_key == "COORDINATOR":
                            current_entry["coordinator"] = str(val).strip()
                        elif out_key == "INVIGILATOR":
                            current_entry["invigilator"] = str(val).strip()
                        break

            current_entry["datetime_str"] = compute_datetime_str(
                current_entry["day"], current_entry["time"]
            )

        if current_entry:
            courses.append(current_entry)

        all_courses.extend(courses)

    return all_courses